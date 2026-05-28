#!/usr/bin/env python3
"""
Aplica precios mayoristas escalonados al Excel del editor masivo de ML.

Uso:
    python3 aplicar_mayoristas.py --input ARCHIVO_INPUT.xlsx [--output ARCHIVO_OUTPUT.xlsx]

Lee la configuración de tiers desde config.json (en la misma carpeta).
"""
from __future__ import annotations
import argparse
import json
import openpyxl
import os
import shutil
import sys
from pathlib import Path

# Mapping fijo de columnas en el formato editor masivo de ML
# (estado al 2026-05-26 - validado con archivos reales)
COLS = {
    'item_id': 2,
    'title': 6,
    'stock': 8,        # STOCK_FLEX
    'price': 9,        # PRICE
    'currency': 10,    # CURRENCY_ID
    # Mayoristas: pares (precio, cantidad)
    'tier_1_price': 11, 'tier_1_qty': 12,
    'tier_2_price': 13, 'tier_2_qty': 14,
    'tier_3_price': 15, 'tier_3_qty': 16,
    'tier_4_price': 17, 'tier_4_qty': 18,
    'tier_5_price': 19, 'tier_5_qty': 20,
}

# Filas 1-5 son encabezados; datos empiezan en fila 6
DATA_START_ROW = 6


def to_float(v):
    """Convierte celda a float, tratando '' y ' ' como None."""
    if v is None or (isinstance(v, str) and v.strip() == ''):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(',', '.'))
    except (ValueError, TypeError):
        return None


def cargar_config(skill_dir: Path) -> dict:
    """Carga config.json desde la carpeta de la skill."""
    cfg_path = skill_dir / 'config.json'
    if not cfg_path.exists():
        raise FileNotFoundError(f'No se encontró {cfg_path}')
    with open(cfg_path, encoding='utf-8') as f:
        cfg = json.load(f)
    tiers = cfg.get('tiers', [])
    if not tiers or len(tiers) > cfg.get('max_tiers_ml', 5):
        raise ValueError(f'tiers debe tener 1-{cfg.get("max_tiers_ml",5)} entradas')
    # Ordenar por min_qty asc
    tiers.sort(key=lambda t: t['min_qty'])
    return cfg


def copiar_input_writable(src: str, dst: str) -> None:
    """
    Copia el archivo de entrada al destino con permisos de escritura.
    `shutil.copy` puede heredar permisos read-only del directorio uploads, así
    que usamos `cat` (en realidad: read+write binario) y luego chmod.
    """
    with open(src, 'rb') as fi, open(dst, 'wb') as fo:
        fo.write(fi.read())
    os.chmod(dst, 0o644)


def aplicar(input_path: str, output_path: str, skill_dir: Path) -> dict:
    """
    Aplica los escalones mayoristas y guarda el archivo modificado.
    Devuelve un dict con métricas de la corrida.
    """
    cfg = cargar_config(skill_dir)
    tiers = cfg['tiers']

    copiar_input_writable(input_path, output_path)

    wb = openpyxl.load_workbook(output_path)
    if 'Publicaciones' not in wb.sheetnames:
        raise ValueError("El archivo no contiene la hoja 'Publicaciones'. ¿Es realmente del editor masivo de ML?")
    ws = wb['Publicaciones']

    # PASO 1: Limpiar TODAS las columnas de mayoristas (cols 11-20)
    celdas_borradas = 0
    for r in range(DATA_START_ROW, ws.max_row + 1):
        for col in range(11, 21):
            if ws.cell(row=r, column=col).value not in (None, '', ' '):
                ws.cell(row=r, column=col).value = None
                celdas_borradas += 1

    # PASO 2: Aplicar escalones según stock
    stats = {
        'celdas_borradas': celdas_borradas,
        'sin_stock': 0,
        'sin_precio': 0,
        'tiers_aplicados': [0] * len(tiers),
        'tier_min_qty': [t['min_qty'] for t in tiers],
        'tier_discount': [t['discount_pct'] for t in tiers],
    }
    sample = []

    for r in range(DATA_START_ROW, ws.max_row + 1):
        stock = to_float(ws.cell(row=r, column=COLS['stock']).value)
        price = to_float(ws.cell(row=r, column=COLS['price']).value)

        if price is None:
            stats['sin_precio'] += 1
            continue
        if stock is None or stock < tiers[0]['min_qty']:
            stats['sin_stock'] += 1
            continue

        for i, tier in enumerate(tiers):
            if stock >= tier['min_qty']:
                price_col = COLS[f'tier_{i+1}_price']
                qty_col = COLS[f'tier_{i+1}_qty']
                tier_price = round(price * (1 - tier['discount_pct'] / 100.0), 2)
                ws.cell(row=r, column=price_col).value = tier_price
                ws.cell(row=r, column=qty_col).value = tier['min_qty']
                stats['tiers_aplicados'][i] += 1

        # Muestra para reporte
        if len(sample) < 6 and stock >= tiers[-1]['min_qty']:
            sample.append({
                'fila': r,
                'item_id': ws.cell(row=r, column=COLS['item_id']).value,
                'titulo': ws.cell(row=r, column=COLS['title']).value,
                'stock': stock,
                'precio': price,
                'tiers': [round(price * (1 - t['discount_pct'] / 100.0), 2) for t in tiers],
            })

    wb.save(output_path)
    stats['sample'] = sample
    stats['output_path'] = output_path
    return stats


def imprimir_resumen(stats: dict) -> None:
    print(f"\n✓ Archivo guardado: {stats['output_path']}")
    print(f"\n=== RESUMEN ===")
    print(f"  Celdas mayoristas limpiadas: {stats['celdas_borradas']}")
    for i, n in enumerate(stats['tiers_aplicados']):
        q = stats['tier_min_qty'][i]
        d = stats['tier_discount'][i]
        print(f"  Mayorista {i+1} (≥{q}u, {d}% off): {n} filas")
    print(f"  Stock<{stats['tier_min_qty'][0]} (vacías): {stats['sin_stock']} filas")
    print(f"  Sin precio (vacías): {stats['sin_precio']} filas")
    if stats['sample']:
        print(f"\n=== MUESTRA (filas con todos los escalones) ===")
        for it in stats['sample']:
            print(f"\n  F{it['fila']:>4} | {it['item_id']} | stock={it['stock']:>3.0f} | precio=${it['precio']:>10,.2f}")
            print(f"           {it['titulo'][:88]}")
            tier_str = "  |  ".join(
                f"M{i+1} (≥{stats['tier_min_qty'][i]}u): ${p:>10,.2f}"
                for i, p in enumerate(it['tiers'])
            )
            print(f"           → {tier_str}")


def main():
    ap = argparse.ArgumentParser(description='Aplica precios mayoristas al editor masivo de ML.')
    ap.add_argument('--input', required=True, help='Archivo Excel del editor masivo de ML')
    ap.add_argument('--output', help='Archivo de salida (default: input_CON_MAYORISTA.xlsx)')
    args = ap.parse_args()

    input_path = args.input
    if not os.path.exists(input_path):
        print(f'ERROR: no existe {input_path}', file=sys.stderr)
        sys.exit(1)

    if args.output:
        output_path = args.output
    else:
        base, ext = os.path.splitext(input_path)
        output_path = f'{base}_CON_MAYORISTA{ext}'

    skill_dir = Path(__file__).resolve().parent
    stats = aplicar(input_path, output_path, skill_dir)
    imprimir_resumen(stats)


if __name__ == '__main__':
    main()
