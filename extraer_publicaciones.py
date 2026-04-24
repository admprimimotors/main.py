"""
extraer_publicaciones.py
========================
Extrae las publicaciones actuales de Mercado Libre y las exporta a Excel.

Te permite elegir si querés:
  - solo las activas (las que están vivas en ML ahora)
  - todas (incluyendo pausadas, cerradas, etc.)

Uso:
    python extraer_publicaciones.py
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import config
from logger import get_logger
from ml import client
from ml.auth import cargar_tokens

log = get_logger(__name__)


# ==========================================================
# Obtención de IDs (con scroll para catálogos grandes)
# ==========================================================
def obtener_ids_publicaciones(user_id: int, solo_activas: bool) -> list[str]:
    """
    Devuelve la lista de IDs de todas las publicaciones del usuario.
    Usa el endpoint de scroll porque soporta >1000 items sin problema.
    """
    ids: list[str] = []
    scroll_id: str | None = None
    endpoint = f"/users/{user_id}/items/search"

    while True:
        params: dict = {"search_type": "scan", "limit": 100}
        if solo_activas:
            params["status"] = "active"
        if scroll_id:
            params["scroll_id"] = scroll_id

        resp = client.get(endpoint, params=params)
        batch = resp.get("results", []) or []
        if not batch:
            break

        ids.extend(batch)
        print(f"  ...obtenidos {len(ids)} IDs hasta ahora")

        scroll_id = resp.get("scroll_id")
        if not scroll_id:
            break

    return ids


# ==========================================================
# Obtención de detalles en lote (multi-get de 20 en 20)
# ==========================================================
def obtener_detalles(item_ids: list[str]) -> list[dict]:
    """
    Consulta /items con multi-get. ML permite hasta 20 items por call.
    """
    detalles: list[dict] = []
    lote = 20

    for i in range(0, len(item_ids), lote):
        batch = item_ids[i:i + lote]
        ids_str = ",".join(batch)
        resp = client.get("/items", params={"ids": ids_str})

        for wrapper in resp:
            if wrapper.get("code") == 200 and wrapper.get("body"):
                detalles.append(wrapper["body"])
            else:
                log.warning(f"No se pudo obtener item: {wrapper}")

        procesados = min(i + lote, len(item_ids))
        print(f"  ...procesados {procesados}/{len(item_ids)}")

    return detalles


# ==========================================================
# Extractor de SKU (primero campo dedicado, después atributos)
# ==========================================================
def extraer_sku(item: dict) -> str:
    """
    El SKU puede estar en:
     1) item.seller_custom_field  (campo clásico)
     2) item.attributes con id = SELLER_SKU  (ML Catalog)
    """
    sku = item.get("seller_custom_field") or ""
    if sku:
        return str(sku)

    for attr in item.get("attributes", []) or []:
        if attr.get("id") == "SELLER_SKU":
            return str(attr.get("value_name") or "")

    return ""


# ==========================================================
# Exportador a Excel
# ==========================================================
def exportar_a_excel(items: list[dict], destino: Path) -> None:
    """Genera un Excel con una fila por publicación y encabezados formateados."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Publicaciones ML"

    encabezados = [
        "Item ID", "SKU", "Título", "Estado", "Condición",
        "Precio", "Moneda", "Stock disponible", "Vendidos",
        "Categoría ID", "Tipo publicación", "Modo envío",
        "Link público", "Thumbnail",
        "Fecha creación", "Última modificación",
    ]
    ws.append(encabezados)

    # Formato del encabezado
    for col_num, _ in enumerate(encabezados, start=1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F3864")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Filas de datos
    for item in items:
        shipping = item.get("shipping") or {}
        fila = [
            item.get("id"),
            extraer_sku(item),
            item.get("title"),
            item.get("status"),
            item.get("condition"),
            item.get("price"),
            item.get("currency_id"),
            item.get("available_quantity"),
            item.get("sold_quantity"),
            item.get("category_id"),
            item.get("listing_type_id"),
            shipping.get("mode"),
            item.get("permalink"),
            item.get("thumbnail"),
            item.get("date_created"),
            item.get("last_updated"),
        ]
        ws.append(fila)

    # Inmovilizar la primera fila
    ws.freeze_panes = "A2"

    # Auto-ajuste de ancho de columnas (con tope)
    for col in ws.columns:
        max_len = 0
        letra = get_column_letter(col[0].column)
        for cell in col:
            val = "" if cell.value is None else str(cell.value)
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[letra].width = min(max(max_len + 2, 10), 60)

    wb.save(destino)


# ==========================================================
# Main interactivo
# ==========================================================
def main() -> None:
    print("=" * 70)
    print("   PRIMI MOTORS — Extracción de publicaciones desde Mercado Libre")
    print("=" * 70)
    print()

    tokens = cargar_tokens()
    if tokens is None:
        print("❌ No hay tokens guardados. Corré primero:")
        print("     python get_initial_token.py")
        return

    print(f"Conectado como user_id = {tokens.user_id}")
    print()
    print("¿Qué publicaciones querés extraer?")
    print("  1) Solo ACTIVAS (las que están vivas en ML ahora)")
    print("  2) TODAS (incluyendo pausadas, cerradas, etc.)")
    print()
    opcion = (input("Elegí 1 o 2 (ENTER = 1): ").strip() or "1")
    solo_activas = opcion != "2"

    print()
    print("🔍 Paso 1/3 — Obteniendo IDs de publicaciones...")
    ids = obtener_ids_publicaciones(tokens.user_id, solo_activas)
    print(f"✓ {len(ids)} publicaciones encontradas.")

    if not ids:
        print("No hay publicaciones para extraer. Saliendo.")
        return

    print()
    print("📦 Paso 2/3 — Descargando detalles de cada publicación...")
    detalles = obtener_detalles(ids)
    print(f"✓ Detalles obtenidos: {len(detalles)}")

    print()
    print("💾 Paso 3/3 — Generando Excel...")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    sufijo = "activas" if solo_activas else "todas"
    nombre = f"publicaciones_ML_{sufijo}_{timestamp}.xlsx"
    destino = config.BASE_DIR / nombre
    exportar_a_excel(detalles, destino)

    print()
    print("=" * 70)
    print("  ✅ LISTO")
    print("=" * 70)
    print(f"  Archivo generado: {nombre}")
    print(f"  Ruta completa:    {destino}")
    print()
    print("Abrí el Excel y fijate:")
    print("  • Qué SKUs tenés cargados (columna SKU)")
    print("  • Publicaciones sin SKU (para corregir después)")
    print("  • Stock actual por producto")
    print("  • Precios vigentes")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n⚠ Cancelado por el usuario.")
    except Exception as e:
        log.exception("Error fatal en extracción")
        print(f"\n❌ Ocurrió un error: {e}")
        print("   Revisá el archivo data/logs/primi_motors.log para más detalle.")
